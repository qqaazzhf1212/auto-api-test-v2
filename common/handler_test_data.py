import random
import re
from openpyxl import load_workbook


# excel操作
def get_test_data_from_excel(file, sheet_name):
    # 1、打开excel
    wb = load_workbook(file, read_only=True)

    # 2、打开表
    sh = wb[sheet_name]
    row = sh.max_row
    column = sh.max_column
    # print(row)
    # print(column)

    # 3、读数据
    data = []
    keys = []
    for i in range(1, column + 1):
        keys.append(sh.cell(1, i).value)
    # print(keys)

    # 循环每一行，组成字典
    for i in range(2, row + 1):
        # print(i)
        temp = {}
        # 循环每一行的列
        for j in range(1, column + 1):
            # print(j)
            # key=keys[j-1]
            # value = sh.cell(n, m).value
            # print(value)
            # temp[key]=value
            temp[keys[j - 1]] = sh.cell(i, j).value
        # 把request，except_data的json数据转换为python对象

        # try:
        # temp['requests'] = json.loads(temp['requests'])
        # temp['expected'] = json.loads(temp['expected'])
        # temp['check_sql'] = json.loads(temp['check_sql'])
        # except json.decoder.JSONDecodeError:
        # raise ValueError('用例数据json格式错误')
        data.append(temp)
    return data


# 随机生成手机号
def gencreate_phone():
    # 第二位数
    phone = ['1', str(random.randint(3, 9))]
    for i in range(9):
        phone.append(str(random.randint(0, 9)))
    return ''.join(phone)


# 正则表达式处理数据，动态替换参数
def replace_args(json_s, obj):
    pass
    # 1、匹配所有的槽位的变量名
    args = re.findall('#(.+?)#', json_s)
    # 2、再到obj中找到对应的属性替换
    for arg in args:
        value = getattr(obj, arg, None)
        if value:
            json_s = json_s.replace('#{}#'.format(arg), str(value))
    return json_s


# if __name__ == '__main__':
#     # 获取excel数据
#     # file = '../test_datas/api_cases.xlsx'
#     # sheet_name = '登陆'
#     #     res = get_test_data_from_excel(file, sheet_name)
#     #     print(res)
#
#     # print(gencreate_phone())
#     class Env:
#         name = 'lh'
#         age = 18
#     s = '{"name":"#name#","age":"#age#"}'
#     res = replace_args(s, Env)
#     print(res)
